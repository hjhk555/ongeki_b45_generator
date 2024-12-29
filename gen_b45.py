from PIL import Image, ImageDraw, ImageFont
from typing import cast
from openpyxl import load_workbook
import yaml

# 配置读取
config_file = 'config.yaml'
def read_config() -> dict:
    with open(config_file, 'r', encoding='utf-8') as file:
        return yaml.safe_load(file)

config = read_config()
pic_dir = config.get('pic_dir')
score_table_filename = config.get('score_table_filename')
b45_filename = config.get('b45_filename')
difficulty_font_color = config.get('difficulty_font_color')
difficulty_background_color = config.get('difficulty_background_color')
font_size_song_name = config.get('font_size_song_name')
font_size_small = config.get('font_size_small')
font_size_medium = config.get('font_size_medium')
font_size_large = config.get('font_size_large')
bg_width_spacing = config.get('bg_width_spacing')
bg_height_spacing = config.get('bg_height_spacing')
elem_width_spacing = config.get('elem_width_spacing')
elem_height_spacing = config.get('elem_height_spacing')
card_width = config.get('card_width')
card_height = config.get('card_height')
card_elem_width_spacing = config.get('card_elem_width_spacing')
card_elem_height_spacing = config.get('card_elem_height_spacing')
difficulty_names = config.get('difficulty_names')
column_img = config.get('column_img')
column_title = config.get('column_title')
column_difficulty = config.get('column_difficulty')
column_level_base = config.get('column_level_base')
column_score = config.get('column_score')
column_new = config.get('column_new')

difficulty_font_color = [(color[0], color[1], color[2]) for color in difficulty_font_color]
difficulty_background_color = [(color[0], color[1], color[2]) for color in difficulty_background_color]

# 字体
font_small = ImageFont.truetype('comic.ttf', size=font_size_small)
font_small_song_name = ImageFont.truetype('msgothic.ttc', size=font_size_song_name)
font_medium = ImageFont.truetype('comic.ttf', size=font_size_medium)
font_large = ImageFont.truetype('comic.ttf', size=font_size_large)
# 成绩表列号

# 成绩类
class Score:
    name: str
    difficulty: int
    base: float
    score: int
    rating: float
    
    def __init__(self, name, difficulty, base, score, rating):
        self.name = name
        self.difficulty = difficulty
        self.base = base
        self.score = score
        self.rating = rating

# 计算rating
def calc_rating(base: float, score: int) -> float:
    if score <= 970000:
        res = base - (970000-score)/17500
    elif score > 970000 and score <= 1000000:
        res = base + (score-970000)/20000
    else:
        # score > 1000000
        res = base + 1.5 + (min(score, 1007500)-1000000)/15000
    return round(res, 3)

# 印刷成绩卡
def draw_score(img_bg: Image.Image, img_draw: ImageDraw.ImageDraw, score: Score, x_offset: int, y_offset: int):
    # 曲绘
    pic_size = card_height-card_elem_height_spacing-font_size_small
    song_pic = cast(Image.Image, Image.open(f'./song_pics/{score.name}.png')).resize((pic_size, pic_size))
    img_bg.paste(song_pic, (x_offset, y_offset))
    # 难度
    img_draw.rectangle([(x_offset, y_offset+pic_size), (x_offset+pic_size, y_offset+card_height)], fill=difficulty_background_color[score.difficulty])
    difficulty_width_spacing = (pic_size-img_draw.textlength(difficulty_names[score.difficulty], font=font_small))/2
    img_draw.text((x_offset+difficulty_width_spacing, y_offset+pic_size), difficulty_names[score.difficulty], font=font_small, fill=difficulty_font_color[score.difficulty])
    # 歌名
    name_width = card_width-pic_size-elem_width_spacing
    name_height = card_height-2*card_elem_height_spacing-font_size_small-font_size_medium
    max_song_name_lines = name_height//font_size_song_name
    line = ''
    line_index = 0
    for char in score.name:
        test_line = line+char
        if img_draw.textlength(test_line, font=font_small_song_name) > name_width:
            img_draw.text((x_offset+pic_size+card_elem_width_spacing, y_offset+line_index*font_size_song_name), line, font=font_small_song_name, fill='black')
            line_index += 1
            if line_index >= max_song_name_lines:
                break
            line = str(char)
        else:
            line = test_line
    if line_index < max_song_name_lines:
        img_draw.text((x_offset+pic_size+card_elem_width_spacing, y_offset+line_index*font_size_song_name), line, font=font_small_song_name, fill='black')
        line_index += 1
    # 打印成绩
    img_draw.text((x_offset+pic_size+card_elem_width_spacing, y_offset+card_height-font_size_small-card_elem_height_spacing-font_size_medium), str(score.score), font=font_medium, fill='black')
    img_draw.text((x_offset+pic_size+card_elem_width_spacing, y_offset+card_height-font_size_small), f'{score.base:.1f} -> {score.rating:.3f}', font=font_small, fill='black')

if __name__ == '__main__':
    # 读取成绩表
    best30 = []
    new15 = []
    score_book = load_workbook(score_table_filename)
    score_sheet = score_book.active
    for row_index in range(2, score_sheet.max_row+1):
        try:
            song_score = int(score_sheet[f'{column_score}{row_index}'].value)
            song_base = float(score_sheet[f'{column_level_base}{row_index}'].value)
            song_name = score_sheet[f'{column_title}{row_index}'].value
            song_difficulty = difficulty_names.index(score_sheet[f'{column_difficulty}{row_index}'].value)
            song_is_new = (score_sheet[f'{column_new}{row_index}'].value == 'NEW')
            # 排除异常成绩
            if song_score == 0:
                continue
            song_score = Score(song_name, song_difficulty, song_base, song_score, calc_rating(song_base, song_score))
            if song_is_new:
                new15.append(song_score)
            else:
                best30.append(song_score)
        except Exception as e:
            continue
    # 排序
    best30.sort(key=lambda score: score.rating, reverse=True)
    new15.sort(key=lambda score: score.rating, reverse=True)
    best30 = best30[:30]
    new15 = new15[:15]
    total_b30 = sum(score.rating for score in best30)
    total_n15 = sum(score.rating for score in new15)
    total_b45 = total_b30+total_n15
    rating_b30 = round(total_b30/30, 3)
    rating_n15 = round(total_n15/15, 3)
    rating_b45 = round(total_b45/45, 3)

    img_background = cast(Image.Image, Image.open('./resources/ongeki_bg.jpg'))
    image_draw = ImageDraw.Draw(img_background)
    bg_width = img_background.size[0]
    bg_height = img_background.size[1]
    total_width = 2*bg_width_spacing+5*card_width+4*elem_width_spacing
    total_height = 2*bg_height_spacing+10*elem_height_spacing+2*font_size_large+9*card_height
    rating_width = image_draw.textlength(f'B45:{rating_b45}', font=font_large)
    rating_width_spacing = (total_width-2*rating_width)/5
    bg_rating_width_spacing = rating_width_spacing*2

    # rating
    image_draw.text((bg_rating_width_spacing, bg_height_spacing), f'B45:{rating_b45:.3f}', font=font_large, fill='black')
    image_draw.text((bg_rating_width_spacing+rating_width+rating_width_spacing, bg_height_spacing), f'B30:{rating_b30:.3f}', font=font_large, fill='black')
    image_draw.text(((total_width-rating_width)/2, bg_height_spacing+7*elem_height_spacing+6*card_height+font_size_large), f'N15:{rating_n15:.3f}', font=font_large, fill='black')
    # b30
    for i in range(len(best30)):
        width_offset = bg_width_spacing+(i%5)*(card_width+elem_width_spacing)
        height_offset = bg_height_spacing+font_size_large+elem_height_spacing+(i//5)*(card_height+elem_height_spacing)
        draw_score(img_background, image_draw, best30[i], width_offset, height_offset)
    # n15
    for i in range(len(new15)):
        width_offset = bg_width_spacing+(i%5)*(card_width+elem_width_spacing)
        height_offset = bg_height_spacing+2*(font_size_large+elem_height_spacing)+(i//5+6)*(card_height+elem_height_spacing)
        draw_score(img_background, image_draw, new15[i], width_offset, height_offset)
    img_background = img_background.crop((0, 0, total_width, total_height))
    img_background.save(b45_filename)
    print(f'成绩已保存为{b45_filename}')

from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.drawing.spreadsheet_drawing import TwoCellAnchor, AnchorMarker
from openpyxl.styles import Alignment, Font
from PIL import Image as PILImage
import sqlite3, os, yaml

# 配置读取
config_file = 'config.yaml'
def read_config() -> dict:
    with open(config_file, 'r', encoding='utf-8') as file:
        return yaml.safe_load(file)

config = read_config()
db_name = config.get('db_name')
table_name_level = config.get('table_name_level')
table_name_new = config.get('table_name_new')
table_name_lun_new = config.get('table_name_lun_new')
pic_dir = config.get('pic_dir')
compress_pic_dir = config.get('compress_pic_dir')
score_table_filename = config.get('score_table_filename')
difficulty_names = config.get('difficulty_names')
title_row = config.get('title_row')
column_width = config.get('column_width')
row_height = config.get('row_height')
difficulty_color = config.get('difficulty_color')
column_img = config.get('column_img')
column_title = config.get('column_title')
column_difficulty = config.get('column_difficulty')
column_level = config.get('column_level')
column_level_base = config.get('column_level_base')
column_score = config.get('column_score')
column_rating = config.get('column_rating')
column_combo = config.get('column_combo')
column_bell = config.get('column_bell')
column_new = config.get('column_new')

white_font = Font(color='ffffff')
difficulty_font = [Font(color=color) for color in difficulty_color]
column_img_index = ord(column_img) - ord('A') + 1

# sql
sql_difficulty_cols = ['bas', 'adv', 'exp', 'mas', 'lun']
sql_read_levels = f'''SELECT id, name, bas, adv, exp, mas, lun FROM {table_name_level}'''
sql_read_new_songs = f'''SELECT id FROM {table_name_new}'''
sql_read_new_lun_songs = f'''SELECT id FROM {table_name_lun_new}'''

# 根据定数判断难度
def resolveDifficulty(detailed_level: float) -> str:
    base = int(detailed_level)
    detailed_level -= base
    res = str(base)
    if detailed_level-0.7>=-1e-6:
        res+='+'
    return res

if __name__ == '__main__':
    # 初始化表格
    wb = Workbook()
    ws = wb.active
    ws.title = "成绩录入"
    ws.append(title_row)
    middle_alignment = Alignment(horizontal='center', vertical='center')
    for col_index in range(len(column_width)):
        ws.column_dimensions[chr(ord('A')+col_index)].width = column_width[col_index]
        ws[f'{chr(ord('A')+col_index)}{1}'].alignment = middle_alignment
    ws.freeze_panes = 'A2'  # 冻结第一行
    ws.auto_filter.ref = f'A1:{chr(ord('A')+len(title_row)-1)}1'    # 启用过滤功能

    # 创建压缩曲绘文件夹
    try:
        os.mkdir(compress_pic_dir)
    except FileExistsError:
        #ignore
        pass

    # 下拉列表
    dv_combo = DataValidation(type='list', formula1='"-,FC,AB"')
    dv_bell = DataValidation(type='list', formula1='"-,FB"')
    ws.add_data_validation(dv_combo)
    ws.add_data_validation(dv_bell)

    # 读取新歌
    db_conn = sqlite3.connect(db_name)
    db_cursor = db_conn.cursor()
    db_cursor.execute(sql_read_new_songs)
    new_songs = set([row[0] for row in db_cursor.fetchall()])
    db_cursor.execute(sql_read_new_lun_songs)
    new_lun_songs = set([row[0] for row in db_cursor.fetchall()])

    # 已有表格数据继承
    score_map = {}
    combo_map = {}
    bell_map = {}
    if os.path.exists(score_table_filename):
        print('读取并继承已有成绩表......')
        workbook = load_workbook(score_table_filename)
        sheet = workbook.active
        for row_index in range(2, sheet.max_row+1):
            song_id = sheet[f'{column_img}{row_index}'].value or sheet[f'{column_title}{row_index}'].value
            song_difficulty = sheet[f'{column_difficulty}{row_index}'].value
            key = f'{song_id}[{song_difficulty}]'
            score = sheet[f'{column_score}{row_index}'].value
            if score is not None:
                score_map[key] = score
            combo = sheet[f'{column_combo}{row_index}'].value
            if combo is not None:
                combo_map[key] = combo
            bell = sheet[f'{column_bell}{row_index}'].value
            if bell is not None:
                bell_map[key] = bell

    img_height = row_height * 1.33
    img_width = ws.column_dimensions[column_img].width * 7.5
    img_spaceing = 30000

    # 读取定数写表格
    db_cursor.execute(sql_read_levels)
    row_index = 1
    current_song = db_cursor.fetchone()
    while current_song is not None:
        song_id = current_song[0]
        song_name = current_song[1]
        song_img_path = f'./{pic_dir}/{song_id}.png'
        compress_img_path = f'./{compress_pic_dir}/{song_id}.jpeg'
        load_img = os.path.exists(song_img_path)
        if load_img:
            # 压缩图片
            img = PILImage.open(song_img_path)
            img.save(compress_img_path, 'JPEG', quality=20)

        for diff_index in range(len(sql_difficulty_cols)):
            level_base = current_song[diff_index+2]
            if level_base is None:
                continue
            level_base = float(level_base)

            row_index+=1
            ws.row_dimensions[row_index].height = row_height
            for col_index in range(len(column_width)):
                ws[f'{chr(ord('A')+col_index)}{row_index}'].alignment = middle_alignment

            if load_img:
                song_img = Image(compress_img_path)
                song_img.height = img_height
                song_img.width = img_width
                song_img.anchor = TwoCellAnchor(editAs='twoCell', _from=AnchorMarker(col=column_img_index-1, row=row_index-1, colOff=img_spaceing, rowOff=img_spaceing), to=AnchorMarker(col=column_img_index, row=row_index, colOff=-img_spaceing, rowOff=-img_spaceing))
                ws.add_image(song_img)

            ws[f'{column_img}{row_index}'] = song_id
            ws[f'{column_title}{row_index}'] = song_name
            ws[f'{column_difficulty}{row_index}'] = difficulty_names[diff_index]
            ws[f'{column_level}{row_index}'] = resolveDifficulty(level_base)
            ws[f'{column_level_base}{row_index}'] = level_base
            ws[f'{column_rating}{row_index}'] = '=MAX(0,${lvl}{row}+MAX(MIN(1007500,${scr}{row})-1000000,0)/15000+MAX(MIN(1000000,${scr}{row})-970000,0)/20000-MAX(970000-${scr}{row},0)/17500)'.format(lvl=column_level_base, scr=column_score, row=row_index)
            if (sql_difficulty_cols[diff_index] != 'lun' and song_id in new_songs) or (sql_difficulty_cols[diff_index] == 'lun' and song_id in new_lun_songs):
                ws[f'{column_new}{row_index}'] = 'NEW'

            ws[f'{column_img}{row_index}'].font = white_font
            ws[f'{column_title}{row_index}'].font = difficulty_font[diff_index]
            ws[f'{column_difficulty}{row_index}'].font = difficulty_font[diff_index]
            ws[f'{column_level}{row_index}'].font = difficulty_font[diff_index]
            ws[f'{column_level_base}{row_index}'].font = difficulty_font[diff_index]

            map_key = f'{song_id}[{difficulty_names[diff_index]}]'
            ws[f'{column_score}{row_index}'] = score_map.pop(map_key, None)
            ws[f'{column_combo}{row_index}'] = combo_map.pop(map_key, None)
            ws[f'{column_bell}{row_index}'] = bell_map.pop(map_key, None)

            dv_combo.add(ws[f'{column_combo}{row_index}'])
            dv_bell.add(ws[f'{column_bell}{row_index}'])

            if (row_index-1)%100 == 0:
                print(f'已载入{row_index-1}张铺面')
        current_song = db_cursor.fetchone()

    print(f'正在保存......')
    wb.save(score_table_filename)
    print(f"已生成{score_table_filename}")
    if len(score_map) >0 or len(combo_map) >0 or len(bell_map) >0:
        print("以下成绩未能继承，可手动填入：")
        maps = [score_map, combo_map, bell_map]
        for map in maps:
            for key in list(map.keys()):
                print(f'{key}: {score_map.pop(key, 0)} {combo_map.pop(key, '')} {bell_map.pop(key, '')}')
